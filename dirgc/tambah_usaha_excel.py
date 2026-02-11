import math
import os

from .settings import DEFAULT_EXCEL_FILE


def normalize_text(value):
    if value is None:
        return ""
    if isinstance(value, float):
        if math.isnan(value):
            return ""
        if value.is_integer():
            return str(int(value))
    return str(value).strip()


def normalize_lat_lon(value, min_value, max_value):
    """
    Normalize latitude/longitude
    Handle case where longitude stored as integer (e.g., 1014423648 -> 101.4423648)
    """
    if value is None:
        return ""
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return ""
    try:
        number = float(value)
    except (TypeError, ValueError):
        return ""
    
    if math.isnan(number):
        return ""
    
    # Special handling for longitude stored as large integer
    # e.g., 1014423648 -> 101.4423648
    if number > 1000 and min_value == -180 and max_value == 180:
        # Kemungkinan format integer tanpa decimal point
        str_num = str(int(number))
        if len(str_num) >= 9:  # e.g., 1014423648
            # Insert decimal after 3rd digit: 101.4423648
            number = float(str_num[:3] + '.' + str_num[3:])
    
    if number < min_value or number > max_value:
        return ""
    
    return str(number)


def resolve_tambah_usaha_excel_path(excel_file):
    """
    Resolve path for Tambah Usaha Excel file
    Default: tambah_usaha.xlsx di folder data/
    """
    if excel_file:
        return os.path.expanduser(excel_file)

    # Coba di folder data/ dulu
    default_file = os.path.join("data", "tambah_usaha.xlsx")
    if os.path.exists(default_file):
        return default_file
    
    # Fallback ke current directory
    fallback_file = "tambah_usaha.xlsx"
    if os.path.exists(fallback_file):
        return fallback_file

    raise FileNotFoundError(
        "Excel file not found. Use --excel-file or place "
        f"tambah_usaha.xlsx in data/ folder or current directory."
    )


def load_tambah_usaha_rows(excel_path):
    """
    Load rows from tambah_usaha.xlsx
    Format: HORIZONTAL (normal Excel table)
    
    PENTING: Kolom wilayah harus berisi NAMA WILAYAH (teks), bukan kode!
    Contoh:
    - Provinsi: "RIAU" atau "Riau" (bukan "14")
    - Kabupaten: "PEKANBARU" (bukan "71")
    - Kecamatan: "TAMPAN" (bukan "070")
    - Kelurahan: "SIMPANG BARU" (bukan "001")
    
    Jika masih berupa kode, sistem akan mencoba mencari dengan kode tersebut.
    """
    path = resolve_tambah_usaha_excel_path(excel_path)

    try:
        import pandas as pd
    except ImportError:
        pd = None

    if pd:
        try:
            df = pd.read_excel(path, dtype=str)
        except Exception:
            df = None
        if df is None:
            pd = None

    if pd:
        rows = []
        
        # Normalize column names
        df.columns = [str(col).strip().lower().replace(" ", "_") for col in df.columns]
        
        # Mapping kolom
        col_map = {
            "nama_usaha": ["nama_usaha", "nama"],
            "alamat": ["alamat_usaha", "alamat"],
            "provinsi": ["provinsi", "prov"],
            "kabupaten": ["kabupaten", "kabupaten_kota", "kab", "kota"],
            "kecamatan": ["kecamatan", "kec"],
            "kelurahan": ["kelurahan", "desa", "kel"],
            "latitude": ["latitude", "lat"],
            "longitude": ["longitude", "long", "lon", "lng"],
        }
        
        def find_column(field_name):
            for possible_name in col_map.get(field_name, []):
                if possible_name in df.columns:
                    return possible_name
            return None
        
        col_nama = find_column("nama_usaha")
        col_alamat = find_column("alamat")
        col_provinsi = find_column("provinsi")
        col_kabupaten = find_column("kabupaten")
        col_kecamatan = find_column("kecamatan")
        col_kelurahan = find_column("kelurahan")
        col_lat = find_column("latitude")
        col_lon = find_column("longitude")
        
        # Iterate through rows
        for _, row in df.iterrows():
            nama_usaha = normalize_text(row[col_nama]) if col_nama else ""
            alamat = normalize_text(row[col_alamat]) if col_alamat else ""
            
            # Ambil nama wilayah sebagai teks (BUKAN kode!)
            provinsi = normalize_text(row[col_provinsi]) if col_provinsi else ""
            kabupaten = normalize_text(row[col_kabupaten]) if col_kabupaten else ""
            kecamatan = normalize_text(row[col_kecamatan]) if col_kecamatan else ""
            kelurahan = normalize_text(row[col_kelurahan]) if col_kelurahan else ""
            
            latitude = normalize_lat_lon(
                row[col_lat] if col_lat else "", -90, 90
            )
            longitude = normalize_lat_lon(
                row[col_lon] if col_lon else "", -180, 180
            )
            
            # Skip jika nama_usaha kosong
            if not nama_usaha:
                continue
            
            record = {
                "idsbr": "",
                "nama_usaha": nama_usaha,
                "alamat": alamat,
                "provinsi": provinsi,
                "kabupaten": kabupaten,
                "kecamatan": kecamatan,
                "kelurahan": kelurahan,
                "latitude": latitude,
                "longitude": longitude,
                "hasil_gc": 2,
            }
            
            rows.append(record)
        
        return rows

    # Fallback to openpyxl
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError(
            "Install pandas or openpyxl to read Excel files."
        ) from exc

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = []
        
        # Read header row
        headers = []
        for col_idx in range(1, sheet.max_column + 1):
            header = normalize_text(sheet.cell(row=1, column=col_idx).value)
            headers.append(header.lower().replace(" ", "_"))
        
        # Mapping kolom
        col_map = {
            "nama_usaha": ["nama_usaha", "nama"],
            "alamat": ["alamat_usaha", "alamat"],
            "provinsi": ["provinsi", "prov"],
            "kabupaten": ["kabupaten", "kabupaten_kota", "kab", "kota"],
            "kecamatan": ["kecamatan", "kec"],
            "kelurahan": ["kelurahan", "desa", "kel"],
            "latitude": ["latitude", "lat"],
            "longitude": ["longitude", "long", "lon", "lng"],
        }
        
        def find_column_index(field_name):
            for possible_name in col_map.get(field_name, []):
                try:
                    return headers.index(possible_name) + 1
                except ValueError:
                    continue
            return None
        
        col_nama = find_column_index("nama_usaha")
        col_alamat = find_column_index("alamat")
        col_provinsi = find_column_index("provinsi")
        col_kabupaten = find_column_index("kabupaten")
        col_kecamatan = find_column_index("kecamatan")
        col_kelurahan = find_column_index("kelurahan")
        col_lat = find_column_index("latitude")
        col_lon = find_column_index("longitude")
        
        # Iterate through data rows
        for row_idx in range(2, sheet.max_row + 1):
            def get_cell_value(col_index):
                if not col_index:
                    return None
                return sheet.cell(row=row_idx, column=col_index).value
            
            nama_usaha = normalize_text(get_cell_value(col_nama))
            alamat = normalize_text(get_cell_value(col_alamat))
            provinsi = normalize_text(get_cell_value(col_provinsi))
            kabupaten = normalize_text(get_cell_value(col_kabupaten))
            kecamatan = normalize_text(get_cell_value(col_kecamatan))
            kelurahan = normalize_text(get_cell_value(col_kelurahan))
            latitude = normalize_lat_lon(get_cell_value(col_lat), -90, 90)
            longitude = normalize_lat_lon(get_cell_value(col_lon), -180, 180)
            
            if not nama_usaha:
                continue
            
            record = {
                "idsbr": "",
                "nama_usaha": nama_usaha,
                "alamat": alamat,
                "provinsi": provinsi,
                "kabupaten": kabupaten,
                "kecamatan": kecamatan,
                "kelurahan": kelurahan,
                "latitude": latitude,
                "longitude": longitude,
                "hasil_gc": 2,
            }
            
            rows.append(record)
        
        return rows
    finally:
        workbook.close()
